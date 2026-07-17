import openpyxl
from pathlib import Path
import shutil
from database import add_penduduk, get_penduduk_by_nik, get_resource_path

class ExcelHandler:
    @staticmethod
    def read_headers(file_path):
        """Read the first row of an Excel sheet to get column headers."""
        if not Path(file_path).exists():
            return []
        
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            sheet = wb.active
            for row in sheet.iter_rows(max_row=1, values_only=True):
                headers = [str(cell).strip() if cell is not None else f"Column {i+1}" for i, cell in enumerate(row)]
                wb.close()
                return headers
        except Exception as e:
            print("Error reading headers:", str(e))
            return []
        return []

    @staticmethod
    def preview_data(file_path, limit=10):
        """Read the first few rows of data (excluding headers) for a preview table."""
        if not Path(file_path).exists():
            return []
        
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            sheet = wb.active
            preview_rows = []
            
            # Skip the first row (header)
            row_iter = sheet.iter_rows(values_only=True)
            try:
                next(row_iter) # skip header
            except StopIteration:
                wb.close()
                return []
                
            count = 0
            for row in row_iter:
                if count >= limit:
                    break
                # convert all values to string or none
                preview_rows.append([str(c) if c is not None else "" for c in row])
                count += 1
                
            wb.close()
            return preview_rows
        except Exception as e:
            print("Error previewing data:", str(e))
            return []

    @staticmethod
    def import_excel(file_path, column_mapping):
        """
        Import Excel data to SQLite.
        column_mapping: dict mapping DB field name -> Excel header name or index
        """
        if not Path(file_path).exists():
            return 0, 0, ["Berkas Excel tidak ditemukan."]

        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            sheet = wb.active
            
            row_iter = sheet.iter_rows(values_only=True)
            headers = []
            try:
                first_row = next(row_iter)
                headers = [str(cell).strip() if cell is not None else "" for cell in first_row]
            except StopIteration:
                wb.close()
                return 0, 0, ["Berkas Excel kosong."]

            # Create a lookup: header name -> column index
            header_indices = {name: idx for idx, name in enumerate(headers)}
            
            success_count = 0
            error_count = 0
            logs = []
            row_num = 1 # Header is row 1
            
            db_fields = [
                "nik", "kk", "nama", "tempat_lahir", "tanggal_lahir", "jenis_kelamin", "agama",
                "status_perkawinan", "pekerjaan", "alamat", "rt", "rw", "desa", "kecamatan",
                "kabupaten", "provinsi", "kewarganegaraan"
            ]

            for row in row_iter:
                row_num += 1
                # Skip completely empty rows
                if all(cell is None for cell in row):
                    continue
                
                # Build data dict
                resident_data = {}
                for field in db_fields:
                    mapped_header = column_mapping.get(field)
                    
                    val = ""
                    if mapped_header in header_indices:
                        col_idx = header_indices[mapped_header]
                        if col_idx < len(row) and row[col_idx] is not None:
                            val_raw = row[col_idx]
                            import datetime
                            if isinstance(val_raw, (datetime.datetime, datetime.date)):
                                val = val_raw.strftime("%Y-%m-%d")
                            elif isinstance(val_raw, float):
                                val = f"{val_raw:.0f}"
                            else:
                                val = str(val_raw).strip()
                    resident_data[field] = val

                # Validasi NIK
                nik = resident_data.get("nik", "")
                if not nik:
                    error_count += 1
                    logs.append(f"Baris {row_num}: Gagal - NIK kosong.")
                    continue
                
                # NIK must be digits and generally 16 characters
                nik_cleaned = "".join(filter(str.isdigit, nik))
                if len(nik_cleaned) != 16:
                    error_count += 1
                    logs.append(f"Baris {row_num}: Gagal - NIK '{nik}' tidak valid (harus 16 digit angka).")
                    continue
                
                resident_data["nik"] = nik_cleaned

                # Default values for empty fields
                if not resident_data.get("kk"):
                    resident_data["kk"] = "0000000000000000"
                if not resident_data.get("nama"):
                    resident_data["nama"] = "TANPA NAMA"
                if not resident_data.get("kewarganegaraan"):
                    resident_data["kewarganegaraan"] = "WNI"

                # Insert into DB
                success, msg = add_penduduk(resident_data)
                if success:
                    success_count += 1
                else:
                    error_count += 1
                    logs.append(f"Baris {row_num} (NIK {nik_cleaned}): Gagal - {msg}")

            wb.close()
            return success_count, error_count, logs

        except Exception as e:
            return 0, 0, [f"Terjadi kesalahan sistem: {str(e)}"]

    @staticmethod
    def generate_template(output_path):
        """Copy an empty Excel template for data import from the provided template."""
        try:
            template_src = get_resource_path("templates/xls/Template_Data_Penduduk.xlsx")
            if not Path(template_src).exists():
                return False, f"File template sumber tidak ditemukan di: {template_src}"
            
            shutil.copy(template_src, output_path)
            return True, "Template berhasil dibuat."
        except Exception as e:
            return False, f"Gagal membuat template: {str(e)}"
